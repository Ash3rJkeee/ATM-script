import openpyxl
from datetime import datetime, timedelta

from request import *
import win32com.client
import os
import gently_stop

# путь к файлу для записи
PATH = settings.path

# с каким файлом надо работать
FILE = f'{PATH}{settings.file}'

# c каким листом надо работать
SHEET_NAME = settings.sheet


def datetime_to_excel_format_date(date):
    """Возвращает дату в формате Excel"""
    return datetime.strftime(date, "%d.%m.%Y")


def datetime_to_excel_format_time(date):
    """Возвращает время в формате Excel"""
    return datetime.strftime(date, "%H:%M:%S")


def rounded_to_4_hour_and_date(mydatetime: datetime):
    """Принимает datetime.
    Если текущее время больше чем на 2 часа, чем последнее контрольное, то
    округляет до следующего контрольного времени.
    Если текущее время больше 22 часов, округляет время до 0 часов, но прибавляет к дате 1 день.
    Возвращает данные в формате datetime  ГГГГ-ММ-ДД ЧЧ:ММ:СС"""

    hours = [0, 4, 8, 12, 16, 20, 0]

    mydatetime_hour = mydatetime.timetuple().tm_hour
    mydatetime_min = mydatetime.timetuple().tm_min / 60

    for hour in hours:
        if mydatetime_hour not in range(hour, hour + 4):
            continue
        else:
            if (mydatetime_hour + mydatetime_min - hour) <= 2:
                return mydatetime.replace(hour=hour, minute=0, second=0, microsecond=0)
            else:
                if mydatetime_hour + mydatetime_min > 22:
                    mydatetime_hour = hours[int(f'{hours.index(hour) + 1}')]
                    return (mydatetime + timedelta(days=1)).replace(hour=mydatetime_hour, minute=0, second=0,
                                                                    microsecond=0)
                else:
                    mydatetime_hour = hours[int(f'{hours.index(hour) + 1}')]
                    return mydatetime.replace(hour=mydatetime_hour, minute=0, second=0, microsecond=0)


def rounded_hour(mydatetime: datetime):
    """Округляет переданное время до целых часов."""
    if mydatetime.timetuple().tm_min == 59:
        mydatetime_hour = mydatetime.timetuple().tm_hour
        return mydatetime.replace(hour=mydatetime_hour+1, minute=0, second=0, microsecond=0)
    else:
        return mydatetime


def fill_param_in_param_list(param: Parameter):
    """Выполняет проверку на актуальность данных.
    Если данные устарелые, возвращает время последнего успешного опроса.
    Если данные актуальные, возвращает """
    if (param is not None) and param.expired:
        return f"нет данных с {param.last_dt}"
    elif param is None:
        return None
    else:
        return param.val


def write_a_row_to_excel(heat_object):
    """Делает запись в эксель файле в строку с текущим mark-kow ТОЛЬКО В ПУСТЫЕ ЯЧЕЙКИ"""
    colls_list = ['K', 'M', 'N', 'O', 'Q', 'T']

    params_list = [
        fill_param_in_param_list(heat_object.t1),
        fill_param_in_param_list(heat_object.t2),
        fill_param_in_param_list(heat_object.p1),
        fill_param_in_param_list(heat_object.p2),
        fill_param_in_param_list(heat_object.g1),
        fill_param_in_param_list(heat_object.gp_calc)
    ]

    if heat_object.gp is not None:                          # проверка на наличие определенного прибором gp
        params_list.append(fill_param_in_param_list(heat_object.gp))
        colls_list.append('T')

    if heat_object.ta is not None:
        params_list.append(fill_param_in_param_list(heat_object.ta))
        colls_list.append('H')

    print(f'Делаю запись значений параметров для {heat_object.name}')
    for i in range(0, len(params_list)):
        cell = ws[f'{colls_list[i]}{mark_row}']
        if cell.value is None:
            cell.value = params_list[i]


# даты в экселе сохранены в datetime
# время в экселе сохранено в виде datetime.time


# имитация даты 03.10.2020 18:28:00
# day = datetime(2020, 10, 3, 0, 0, 0).date()
# hour = datetime(year=1, day=1, month=1, hour=12, minute=11)
# rounded_to_4_hour = rounded_to_4_hour_and_date(hour)

rounded_to_4_hour = rounded_to_4_hour_and_date(datetime.now())
day = rounded_to_4_hour.date()

start = datetime.now()

# if day > datetime(2020, 10, 9, 0, 0, 0).date():
#     print("Период ознакомления истек. Это нужно, чтобы исключить соблазн использовать не проверенный скрипт в работе.")
#     input("Нажми Enter для завершения работы.")
#     raise PermissionError("Период ознакомления истек")

print(f"Дата {day}")
print(f"Данные будут записаны в блок времени {rounded_to_4_hour.time()}")


print('Пересчет значений формул в файле.')
# запуск пересчета значений формул файла записи

Excel = win32com.client.Dispatch('Excel.Application')

count_copies = Excel.Workbooks.Count

try:
    wb = Excel.Workbooks.Open(FILE)
except:
    print(f'Не удается найти файл "{settings.file}", указанный в настройках.')
    if count_copies == 0:
        Excel.Quit()
    gently_stop.gently_stop_program(code=1)

try:
    ws = wb.Sheets['2020']  # pywintypes.com_error
except:
    print(f'Не удается найти лист "{settings.sheet}" в файле "{settings.file}".')
    if count_copies == 0:
        Excel.Quit()
    gently_stop.gently_stop_program(code=1)

ws.Calculate

wb.Save()
wb.Close()

if count_copies == 0:
    Excel.Quit()
print('Пересчет успешно выполнен. Файл закрыт.')

finish = datetime.now()
print(f'Со старта прошло {(finish - start).seconds} сек')

print('Открываю файл для поиска строки записи.')

# открытие файла для поиска позиции записи
try:
    wb = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
except FileNotFoundError:
    print(f'Не удается найти файл "{settings.file}", указанный в настройках.')
    gently_stop.gently_stop_program(code=1)
except KeyError:
    print(f'Не удается найти лист "{settings.sheet}" в файле "{settings.file}".')
    wb.close()
    gently_stop.gently_stop_program(code=1)

mark_row = 0

# todo Разобраться в причине медленного перебора строк в файле. Причем чем больше строк перебрано, тем дольше поиск.

# ищем блок с нужной датой и временем и сохраняем номер его строки
try:
    print('Начинаю перебор строк.')
    i = 1
    while True:
        i += 1                             # в поиске игнорируем шапку

        if i % 100 == 0:
            print(f'Проверяю {i} строку....')
            finish = datetime.now()
            print(f'Со старта прошло {(finish - start).seconds} сек.')

        cell_date = ws[f'C{i}'].value
        cell_hour = ws[f'D{i}'].value

        if cell_date is None:
            print("Дата не найдена.")
            gently_stop.gently_stop_program(code=1)

        if cell_date.date() == day:
            # так как в таблице дата может оказаться в двух форматах (видимо из-за настроек локали оператора, который
            # делал заготовки на даты.
            if (rounded_hour(cell_hour).time() == rounded_to_4_hour.time()) or (cell_hour == rounded_to_4_hour.time()):
                break

    mark_row = i

    finish = datetime.now()
    print(f'Прошло {(finish - start).seconds} сек')

    print(f'Запись будет сделана начиная с {mark_row} строки.')

    wb.close()

    print('Закрываю файл после поиска записи.')

    print("Открываю файл для записи")
    # повторно открываем файл для записи данных
    wb = openpyxl.load_workbook(FILE)
    ws = wb[SHEET_NAME]

    names_of_epmty_sourses = [
        'hf_rts_150',
        "hf_kts_gornaya",
        "sf_mk_polet"
    ]

    for i in range(len(souses_list)):
        if souses_list[i].name not in names_of_epmty_sourses:           # проверка на не подключенные к АТМ источники
            write_a_row_to_excel(souses_list[i])
        mark_row += 1

    wb.save(FILE)
    wb.close()

    print("Запись успешно сделана. Файл закрыт.")
    print("ВНИМАНИЕ. ДАННЫМ ПОКА ЧТО НЕЛЬЗЯ ДОВЕРЯТЬ. НЕ ДЛЯ ИСПОЛЬЗОВАНИЯ В РАБОТЕ!")

    gently_stop.gently_stop_program(code=0)

except AttributeError:
    print()
    print()
    print("   ++ Нет кэшированных результатов вычисления формул.  ++"
          "   ++ Скрипт не смог запустить пересчет формул самостоятельно.  ++"
          "   ++ Откройте файл в Excel и пересохраните.  ++   ")
    print("   ++ Затем запустите скрипт заново  ++   ")
    print()
    print()
    gently_stop.gently_stop_program(code=1)
